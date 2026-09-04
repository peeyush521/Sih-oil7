from fastapi import FastAPI, HTTPException, UploadFile, File, Depends

from fastapi.middleware.cors import CORSMiddleware

from fastapi.staticfiles import StaticFiles

from fastapi.responses import FileResponse, JSONResponse

from pydantic import BaseModel

from dotenv import load_dotenv

import io

import os

import json
import threading

import datetime

from pathlib import Path



# Load .env file

load_dotenv()



from fastapi.responses import StreamingResponse

from nlp_engine import get_nlp_engine

from graph_engine import get_graph_engine

from risk_engine import calculate_risk, get_intervention_recommendation

from classification_engine import get_classification_engine

from llm_engine import generate_llm_explanation, GEMINI_AVAILABLE
try:
    from google import genai as google_genai
except ImportError:
    google_genai = None

from data_loader import load_industrial_dataset
from mongo_persistence import save_reports, load_reports, save_audit_log, get_audit_logs

from alerts import send_precursor_alert
from rag_engine import generate_rag_analysis
from prediction_engine import predict_next_incidents, get_prediction_summary
from image_detector import detect_hazards_from_image, get_detection_summary
from realtime_alerts import create_precursor_alert, get_alert_history as get_alert_history_rt, send_sms_alert, alert_history
from gnn_engine import get_gnn_engine
from anomaly_detector import get_anomaly_detector
from xai_engine import get_xai_engine
from fatigue_engine import get_fatigue_engine
from auth import signup, login, get_me, get_current_user, SignupRequest, LoginRequest
from language_utils import detect_language, translate_hinglish, get_language_info


class ActionUpdateRequest(BaseModel):
    report_id: str
    action: str  # "complete" or "delay"
    note: str = ""


app = FastAPI(title="SIF Precursor Intelligence System", version="1.0.0")



app.add_middleware(

    CORSMiddleware,

    allow_origins=["*"],

    allow_credentials=True,

    allow_methods=["*"],

    allow_headers=["*"],

)



# Load the real dataset

real_dataset = load_industrial_dataset()

dataset_index = 0



all_processed_reports = []



# â”€â”€ Persistence â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

STATE_FILE = "api_state.json"



def save_state():
    """Persist to MongoDB with JSON fallback."""
    save_reports(all_processed_reports)


def load_persisted_state():
    """Load from MongoDB with JSON fallback. Also rebuilds the knowledge graph."""
    global dataset_index, all_processed_reports
    reports = load_reports()
    if reports:
        all_processed_reports = reports
        dataset_index = len(all_processed_reports)
        print(f"[persist] Restored {len(all_processed_reports)} reports")
        
        # Rebuild knowledge graph from persisted reports
        graph = get_graph_engine()
        nlp = get_nlp_engine()
        for r in all_processed_reports:
            report_data = r.get("report", {})
            entities = r.get("extracted_entities", {})
            if report_data.get("id") and entities:
                embedding = nlp.get_embedding(report_data.get("text", ""))
                graph.add_report(
                    report_data["id"],
                    report_data.get("date", ""),
                    report_data.get("text", ""),
                    entities,
                    embedding,
                    report_data.get("action_status", "Closed")
                )
        print(f"[graph] Rebuilt graph with {graph.graph.number_of_nodes()} nodes, {graph.graph.number_of_edges()} edges")


load_persisted_state()



# â”€â”€ Models â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

class ProcessResponse(BaseModel):

    report: dict

    report_class: str

    classification: dict = {}

    extracted_entities: dict

    risk_data: dict

    llm_explanation: str

    interventions: list

    is_precursor: bool

    rag_analysis: dict = None



class CustomReportRequest(BaseModel):

    text: str



# â”€â”€ Endpoints â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/api/reset")

def reset_state():

    global dataset_index, all_processed_reports

    dataset_index = 0

    all_processed_reports = []

    get_graph_engine().reset()

    if os.path.exists(STATE_FILE):

        os.remove(STATE_FILE)

    return {"status": "reset_complete"}



@app.get("/api/state")

def get_state():

    return {

        "total_reports": len(all_processed_reports),

        "reports": all_processed_reports,

        "is_complete": dataset_index >= len(real_dataset)

    }



@app.post("/api/process_next", response_model=ProcessResponse)

def process_next_report():

    global dataset_index



    if dataset_index >= len(real_dataset):

        raise HTTPException(status_code=400, detail="No more reports in dataset")



    raw_record = real_dataset[dataset_index]



    report = {

        "id": f"RPT-REAL-{dataset_index+1:03d}",

        "date": raw_record["Data"],

        "text": raw_record["Description"],

        "location": raw_record["Local"],

        "action_status": raw_record.get("Action Status", "Closed")

    }



    dataset_index += 1



    try:

        result = _process_report(report)

        all_processed_reports.append(result)

        save_state()

        return result

    except Exception as e:

        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")



@app.post("/api/submit_report", response_model=ProcessResponse)

def submit_custom_report(request: CustomReportRequest):

    import datetime



    if not request.text or not request.text.strip():

        raise HTTPException(status_code=400, detail="Report text cannot be empty")



    report_id = f"RPT-CUS-{len(all_processed_reports)+1:03d}"

    date_str = datetime.date.today().isoformat()



    # Detect and translate Hindi/Hinglish
    lang_info = get_language_info(request.text)
    translated_text, was_translated = translate_hinglish(request.text)

    report = {

        "id": report_id,

        "date": date_str,

        "text": translated_text,

        "location": "Custom_Input_Area",

        "language": lang_info,

        "original_text": request.text

    }



    try:

        result = _process_report(report)

        all_processed_reports.append(result)

        save_state()

        return result

    except Exception as e:

        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")



def _fast_explanation(risk_data, equipment_list, related_reports):
    """Generate a fast rule-based explanation (no API call)."""
    score = risk_data.get("score", 0)
    trajectory = risk_data.get("trajectory", "STABLE")
    deltas = risk_data.get("deltas", {})
    equip = equipment_list[0] if equipment_list else "this area"
    n_related = len(related_reports)
    
    parts = []
    if score >= 70:
        parts.append(f"Risk is elevated for {equip}. We detected {n_related} related events recently.")
    elif score >= 40:
        parts.append(f"Moderate risk for {equip}. {n_related} related events found.")
    else:
        parts.append(f"Low risk for {equip}. Routine monitoring recommended.")
    
    if deltas:
        top = sorted(deltas.items(), key=lambda x: x[1], reverse=True)[:2]
        contrib = ", ".join(f"{k} (+{v})" for k, v in top)
        parts.append(f"Top contributors: {contrib}.")
    
    parts.append(f"The precursor score is {score}.")
    
    if score >= 70:
        parts.append("Monitor closely and consider preventive action.")
    elif score >= 40:
        parts.append("Continue monitoring for changes.")
    else:
        parts.append("No immediate action required.")
    
    return " ".join(parts)


def _generate_llm_in_background(report_id, risk_data, equipment_list, related_reports):
    """Background thread: generates LLM explanation and stores it."""
    try:
        explanation = generate_llm_explanation(risk_data, equipment_list, related_reports)
        # Find and update the stored report
        for r in all_processed_reports:
            if r["report"]["id"] == report_id:
                r["llm_explanation"] = explanation
                break
    except Exception:
        pass


def _process_report(report: dict) -> dict:
    """Core processing pipeline - optimized for speed."""
    nlp = get_nlp_engine()
    graph = get_graph_engine()

    extracted_entities = nlp.extract_entities(report["text"])

    # Inject structured location
    if report.get("location") and report["location"] not in extracted_entities["locations"]:
        extracted_entities["locations"].append(report["location"])

    embedding = nlp.get_embedding(report["text"])

    # Classification
    classifier = get_classification_engine()
    classification = classifier.classify_with_confidence(report["text"])
    report_class = classification["class"]

    graph.add_report(
        report["id"], report["date"], report["text"],
        extracted_entities, embedding, report.get("action_status", "Closed")
    )

    related_reports = graph.get_related_reports(report["id"], nlp)
    risk_data = calculate_risk(report["id"], graph.reports[report["id"]], related_reports)
    interventions = get_intervention_recommendation(risk_data, extracted_entities)

    is_precursor = risk_data["score"] >= 70

    # Fast template explanation (no API call)
    llm_explanation = _fast_explanation(risk_data, extracted_entities.get("equipment", []), related_reports)

    result = {
        "report": report,
        "report_class": report_class,
        "classification": classification,
        "extracted_entities": extracted_entities,
        "risk_data": risk_data,
        "llm_explanation": llm_explanation,
        "interventions": interventions,
        "is_precursor": is_precursor,
        "language": report.get("language", get_language_info(report.get("original_text", report.get("text", ""))))
    }

    # Generate real LLM explanation in background thread
    threading.Thread(
        target=_generate_llm_in_background,
        args=(report["id"], risk_data, extracted_entities.get("equipment", []), related_reports),
        daemon=True
    ).start()

    # Store alert in history AND send email if precursor detected
    if is_precursor:
        create_precursor_alert(report, risk_data, extracted_entities)
        send_precursor_alert(report, risk_data, extracted_entities)

    return result
@app.get("/api/graph_data")

def get_graph_data():

    graph_engine = get_graph_engine()

    elements = []



    for node, data in graph_engine.graph.nodes(data=True):

        elements.append({

            "data": {

                "id": node,

                "label": node,

                "type": data.get("type", "Unknown")

            }

        })



    for u, v, data in graph_engine.graph.edges(data=True):

        elements.append({

            "data": {

                "source": u,

                "target": v,

                "label": data.get("relation", "")

            }

        })



    return elements



@app.post("/api/simulate")

def simulate_intervention(intervention_type: str):

    if not all_processed_reports:

        return {"risk_score": 0, "trajectory": "STABLE"}



    latest_risk = all_processed_reports[-1]["risk_data"]["score"]

    if intervention_type == "delay":
        return {"risk_score": min(latest_risk + 10, 100), "trajectory": "ESCALATING"}
    elif intervention_type == "resolve_action":
        return {"risk_score": max(latest_risk - 25, 10), "trajectory": "DECREASING"}
    elif intervention_type == "inspect":
        return {"risk_score": max(latest_risk - 36, 10), "trajectory": "DECREASING"}
    elif intervention_type == "replace":
        return {"risk_score": max(latest_risk - 69, 5), "trajectory": "DECREASING"}
    return {"risk_score": latest_risk, "trajectory": "STABLE"}



@app.post("/api/report_action")
def report_action(req: ActionUpdateRequest):
    """Mark a report as completed or delayed with a response note."""
    target = None
    for r in all_processed_reports:
        if r["report"]["id"] == req.report_id:
            target = r
            break
    if not target:
        raise HTTPException(status_code=404, detail="Report not found")

    if req.action == "complete":
        target["report"]["action_status"] = "Closed"
        target["response"] = {"status": "completed", "note": req.note, "timestamp": datetime.datetime.now().isoformat()}
        # Recalculate risk after action
        nlp = get_nlp_engine()
        graph = get_graph_engine()
        related = graph.get_related_reports(req.report_id, nlp)
        new_risk = calculate_risk(req.report_id, graph.reports[req.report_id], related)
        target["risk_data"] = new_risk
        target["is_precursor"] = new_risk["score"] >= 70
    elif req.action == "delay":
        target["report"]["action_status"] = "Open"
        target["response"] = {"status": "delayed", "note": req.note, "timestamp": datetime.datetime.now().isoformat()}
        # Risk increases when delayed
        new_score = min(target["risk_data"]["score"] + 10, 100)
        target["risk_data"]["score"] = new_score
        target["risk_data"]["trajectory"] = "ESCALATING"
        target["is_precursor"] = new_score >= 70
    else:
        raise HTTPException(status_code=400, detail="Action must be 'complete' or 'delay'")

    save_state()
    return {
        "report_id": req.report_id,
        "action": req.action,
        "response": target["response"],
        "risk_data": target["risk_data"],
        "is_precursor": target["is_precursor"],
        "message": f"Report {req.report_id} marked as {req.action}. " + (
            "Risk decreased after corrective action." if req.action == "complete" else
            "Risk increased — action was delayed."
        )
    }



class TranslateRequest(BaseModel):
    text: str


@app.post("/api/translate")
def translate_report(req: TranslateRequest):
    """Translate Hinglish/Hindi report to English for non-Hindi speakers."""
    text = req.text.strip()
    if not text:
        raise HTTPException(status_code=400, detail="No text provided")

    lang_info = get_language_info(text)

    # Try Gemini for natural translation
    gemini_translation = None
    if lang_info["is_multilingual"] and google_genai and os.getenv("GEMINI_API_KEY"):
        try:
            client = google_genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
            prompt = "Translate this Hindi/Hinglish safety report to clear, professional English. Keep the technical safety terminology accurate. Only output the translation, nothing else. Text: " + text
            resp = client.models.generate_content(model="gemini-2.5-flash", contents=prompt)
            gemini_translation = resp.text.strip() if resp.text else None
        except Exception as e:
            import sys
            print(f"Gemini translate error: {type(e).__name__}: {e}", file=sys.stderr, flush=True)

    # Fallback to dictionary-based translation
    dict_translation, _ = translate_hinglish(text)

    final_translation = gemini_translation or dict_translation

    return {
        "original": text,
        "translation": final_translation,
        "method": "Gemini AI" if gemini_translation else "Dictionary",
        "source_language": lang_info["detected_language"],
    }


@app.get("/api/export")

def export_reports():

    """Export all processed reports as JSON for offline review."""

    if not all_processed_reports:

        raise HTTPException(status_code=400, detail="No reports to export")



    export_data = {

        "system": "SIF Precursor Intelligence System",

        "total_reports": len(all_processed_reports),

        "precursors_detected": sum(1 for r in all_processed_reports if r.get("is_precursor")),

        "reports": all_processed_reports

    }



    return JSONResponse(

        content=export_data,

        headers={"Content-Disposition": "attachment; filename=sif_report_export.json"}

    )



@app.get("/api/analytics")

def get_analytics():

    """Return analytics data for the analytics dashboard."""

    if not all_processed_reports:

        return {

            "risk_trend": [],

            "location_distribution": {},

            "equipment_frequency": {},

            "hazard_distribution": {},

            "severity_over_time": [],

            "total_reports": 0,

            "precursors_detected": 0

        }



    # Risk trend over time

    risk_trend = []

    for r in all_processed_reports:

        risk_trend.append({

            "id": r["report"].get("id", ""),

            "date": r["report"].get("date", ""),

            "score": r["risk_data"]["score"],

            "trajectory": r["risk_data"]["trajectory"],

            "is_precursor": r.get("is_precursor", False)

        })



    # Location distribution

    location_dist = {}

    for r in all_processed_reports:

        for loc in r["extracted_entities"].get("locations", []):

            location_dist[loc] = location_dist.get(loc, 0) + 1



    # Equipment frequency

    equip_freq = {}

    for r in all_processed_reports:

        for eq in r["extracted_entities"].get("equipment", []):

            equip_freq[eq] = equip_freq.get(eq, 0) + 1



    # Hazard distribution

    hazard_dist = {}

    for r in all_processed_reports:

        for haz in r["extracted_entities"].get("hazards", []):

            hazard_dist[haz] = hazard_dist.get(haz, 0) + 1



    # Severity over time

    severity_over_time = []

    for r in all_processed_reports:

        severity_over_time.append({

            "id": r["report"].get("id", ""),

            "date": r["report"].get("date", ""),

            "severity": r["extracted_entities"].get("severity", 1),

            "score": r["risk_data"]["score"]

        })



    return {

        "risk_trend": risk_trend,

        "location_distribution": location_dist,

        "equipment_frequency": equip_freq,

        "hazard_distribution": hazard_dist,

        "severity_over_time": severity_over_time,

        "total_reports": len(all_processed_reports),

        "precursors_detected": sum(1 for r in all_processed_reports if r.get("is_precursor"))

    }



@app.get("/api/export/pdf")

def export_pdf():

    """Generate and download a PDF report of all processed incidents."""

    if not all_processed_reports:

        raise HTTPException(status_code=400, detail="No reports to export")



    try:

        from pdf_generator import generate_pdf_report

        pdf_path = generate_pdf_report(all_processed_reports)

        return FileResponse(

            pdf_path,

            media_type="application/pdf",

            filename=f"sif_report_{datetime.datetime.now().strftime('%Y%m%d')}.pdf"

        )

    except RuntimeError as e:

        raise HTTPException(status_code=500, detail=str(e))



@app.get("/api/health")

def health_check():

    return {

        "status": "healthy",

        "reports_processed": len(all_processed_reports),

        "dataset_remaining": len(real_dataset) - dataset_index,

        "gemini_configured": bool(os.environ.get("GEMINI_API_KEY", ""))

    }



# â”€â”€ Serve Frontend â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# Try to serve the React build first; fall back to static folder


# Add these imports at top of main.py after existing imports
# from fastapi import UploadFile, File
# import io

# Add these endpoints before the static file mount

@app.post("/api/upload_report")
async def upload_report(file: UploadFile = File(...)):
    """Upload a PDF, Excel, or text file containing safety reports."""
    content = await file.read()
    filename = file.filename.lower()
    
    reports_text = []
    
    if filename.endswith(".pdf"):
        try:
            import PyPDF2
            reader = PyPDF2.PdfReader(io.BytesIO(content))
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    reports_text.append(text)
        except ImportError:
            raise HTTPException(status_code=500, detail="PyPDF2 not installed. Run: pip install PyPDF2")
    
    elif filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            desc_idx = headers.index("Description") if "Description" in headers else 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[desc_idx]:
                    reports_text.append(str(row[desc_idx]))
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    
    elif filename.endswith(".csv"):
        import csv
        reader = csv.reader(io.StringIO(content.decode("utf-8").split(chr(10)+chr(10))))
        headers = next(reader, None)
        desc_idx = headers.index("Description") if headers and "Description" in headers else 0
        for row in reader:
            if row and len(row) > desc_idx:
                reports_text.append(row[desc_idx])
    
    elif filename.endswith(".txt"):
        reports_text = content.decode("utf-8").split(chr(10)+chr(10))
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use PDF, Excel, CSV, or TXT.")
    
    if not reports_text:
        raise HTTPException(status_code=400, detail="No report text found in file.")
    
    # Process each extracted report
    results = []
    for text in reports_text[:20]:  # Limit to 20 reports
        if len(text.strip()) < 10:
            continue
        report_id = f"RPT-UPLOADED-{len(all_processed_reports)+1:03d}"
        import datetime
        report = {
            "id": report_id,
            "date": datetime.datetime.now().isoformat(),
            "text": text.strip()[:500],
            "location": "Uploaded_Document"
        }
        try:
            result = _process_report(report)
            all_processed_reports.append(result)
            results.append(result)
        except Exception as e:
            continue
    
    save_state()
    return {"processed": len(results), "reports": results}

@app.get("/api/duplicates")
def get_duplicates():
    """Detect potential duplicate/similar reports."""
    if len(all_processed_reports) < 2:
        return {"duplicates": [], "count": 0}
    
    nlp = get_nlp_engine()
    reports_for_check = [{"id": r["report"]["id"], "text": r["report"]["text"]} for r in all_processed_reports]
    dupes = nlp.detect_duplicates(reports_for_check, threshold=0.75)
    
    return {"duplicates": dupes, "count": len(dupes)}

@app.get("/api/precursor_patterns")
def get_precursor_patterns():
    """Analyze advanced precursor patterns across all reports."""
    if not all_processed_reports:
        return {"patterns": []}
    
    patterns = []
    
    # 1. Same-equipment recurrence
    equip_map = {}
    for r in all_processed_reports:
        for eq in r["extracted_entities"].get("equipment", []):
            if eq not in equip_map:
                equip_map[eq] = []
            equip_map[eq].append(r["report"]["id"])
    for eq, rids in equip_map.items():
        if len(rids) >= 3:
            patterns.append({"type": "Equipment Recurrence", "equipment": eq, "count": len(rids), "reports": rids, "severity": "HIGH"})
    
    # 2. Location clustering
    loc_map = {}
    for r in all_processed_reports:
        for loc in r["extracted_entities"].get("locations", []):
            if loc not in loc_map:
                loc_map[loc] = []
            loc_map[loc].append(r["report"]["id"])
    for loc, rids in loc_map.items():
        if len(rids) >= 4:
            patterns.append({"type": "Location Clustering", "location": loc, "count": len(rids), "reports": rids, "severity": "MEDIUM"})
    
    # 3. Unresolved corrective actions
    open_reports = [r for r in all_processed_reports if r["report"].get("action_status") == "Open"]
    if len(open_reports) >= 3:
        patterns.append({"type": "Unresolved Actions", "count": len(open_reports), "reports": [r["report"]["id"] for r in open_reports], "severity": "HIGH"})
    
    # 4. Severity escalation trend
    recent = all_processed_reports[-5:] if len(all_processed_reports) >= 5 else all_processed_reports
    severities = [r["extracted_entities"].get("severity", 1) for r in recent]
    if len(severities) >= 3 and severities[-1] > severities[0]:
        patterns.append({"type": "Severity Escalation", "trend": "increasing", "recent_severities": severities, "severity": "HIGH"})
    
    return {"patterns": patterns, "total_patterns": len(patterns)}

# --- RAG Root Cause Analysis Endpoint ---
@app.get("/api/rag_analysis/{report_id}")
def get_rag_analysis(report_id: str):
    report_data = None
    for r in all_processed_reports:
        if r["report"]["id"] == report_id:
            report_data = r
            break
    if not report_data:
        raise HTTPException(status_code=404, detail="Report not found")
    nlp = get_nlp_engine()
    graph = get_graph_engine()
    related = graph.get_related_reports(report_id, nlp)
    analysis = generate_rag_analysis(
        report_data["report"]["text"],
        report_data.get("classification", {"class": report_data.get("report_class", "Unknown"), "confidence": 0}),
        report_data["extracted_entities"], report_data["risk_data"], related, graph
    )
    return analysis

# --- Time-Series Prediction Endpoint ---
@app.get("/api/predictions")
def get_predictions():
    preds = predict_next_incidents(all_processed_reports)
    summary = get_prediction_summary(preds["predictions"])
    preds["summary"] = summary
    return preds

# --- Image Hazard Detection Endpoint ---
@app.post("/api/detect_hazards")
async def detect_image_hazards(file: UploadFile = File(...)):
    content_bytes = await file.read()
    if len(content_bytes) > 10000000:
        raise HTTPException(status_code=400, detail="File too large")
    result = detect_hazards_from_image(content_bytes, file.filename or "unknown.jpg")
    result["summary"] = get_detection_summary(result)
    return result

# --- Alert History Endpoint ---
@app.get("/api/alerts")
def get_alerts():
    return get_alert_history_rt()

# --- Alert Acknowledge Endpoint ---
@app.post("/api/alerts/{alert_index}/acknowledge")
def acknowledge_alert(alert_index: int):
    if 0 <= alert_index < len(alert_history):
        alert_history[alert_index]['acknowledged'] = True
        alert_history[alert_index]['acknowledged_at'] = datetime.now().isoformat()
        return {"status": "acknowledged", "alert_index": alert_index}
    raise HTTPException(status_code=404, detail="Alert not found")

# --- Alert Dismiss Endpoint ---
@app.post("/api/alerts/{alert_index}/dismiss")
def dismiss_alert(alert_index: int):
    if 0 <= alert_index < len(alert_history):
        alert_history[alert_index]['dismissed'] = True
        alert_history[alert_index]['dismissed_at'] = datetime.now().isoformat()
        return {"status": "dismissed", "alert_index": alert_index}
    raise HTTPException(status_code=404, detail="Alert not found")

# --- GNN Precursor Detection Endpoint ---
@app.get("/api/gnn_analysis")
def get_gnn_analysis():
    gnn = get_gnn_engine()
    gnn.build_graph(all_processed_reports)
    return gnn.get_gnn_analysis()

# --- Anomaly Detection Endpoint ---
@app.get("/api/anomalies")
def get_anomalies():
    detector = get_anomaly_detector()
    detector.fit(all_processed_reports)
    return detector.detect_anomalies(all_processed_reports)

# --- XAI Explainability Endpoint ---
@app.get("/api/xai/{report_id}")
def get_xai_explanation(report_id: str):
    report_data = None
    for r in all_processed_reports:
        if r["report"]["id"] == report_id:
            report_data = r
            break
    if not report_data:
        raise HTTPException(status_code=404, detail="Report not found")
    xai = get_xai_engine()
    nlp = get_nlp_engine()
    classifier = get_classification_engine()
    classification = report_data.get("classification", {})
    risk_explanation = xai.explain_risk_score(report_data["risk_data"], report_data["extracted_entities"], [])
    classification_explanation = xai.explain_classification(report_data["report"]["text"], classification, nlp, classifier)
    novelty_explanation = None
    if classification.get("is_novel"):
        novelty_explanation = xai.explain_novel_hazard(report_data["report"]["text"], classification, report_data["extracted_entities"])
    return {
        "risk_explanation": risk_explanation,
        "classification_explanation": classification_explanation,
        "novelty_explanation": novelty_explanation,
        "report_id": report_id,
        "report_text": report_data["report"]["text"],
    }

# --- Fatigue Risk Endpoint ---
@app.get("/api/fatigue")
def get_fatigue_analysis():
    engine = get_fatigue_engine()
    return engine.get_crew_fatigue_dashboard(all_processed_reports)

# --- Benchmark endpoint ---
@app.get("/api/benchmark")
def get_benchmark():
    """Return ML benchmark metrics and comparison with baselines."""
    classifier = get_classification_engine()
    metrics = classifier.get_metrics()
    
    # Load pre-computed benchmark data if available
    benchmark_file = "benchmark_data.json"
    if os.path.exists(benchmark_file):
        with open(benchmark_file, "r") as f:
            benchmark = json.load(f)
    else:
        benchmark = {}
    
    return {
        "classifier_metrics": metrics,
        "benchmark": benchmark,
        "dataset_info": {
            "total_reports": len(real_dataset),
            "reports_processed": len(all_processed_reports),
            "dataset_remaining": len(real_dataset) - dataset_index,
        },
        "nlp_info": {
            "equipment_keywords": len(get_nlp_engine().equipment_keywords),
            "hazard_keywords": len(get_nlp_engine().hazard_keywords),
            "location_keywords": len(get_nlp_engine().location_keywords),
            "total_domain_terms": len(get_nlp_engine().equipment_keywords) + len(get_nlp_engine().hazard_keywords) + len(get_nlp_engine().location_keywords),
        }
    }



# --- Auth endpoints ---
@app.post("/api/auth/signup")
async def auth_signup(req: SignupRequest):
    return await signup(req)

@app.post("/api/auth/login")
async def auth_login(req: LoginRequest):
    return await login(req)

@app.get("/api/auth/me")
async def auth_me(user=Depends(get_current_user)):
    return {"user": user}

# --- Confusion Matrix endpoint ---
@app.get("/api/confusion_matrix")
def get_confusion_matrix():
    """Return confusion matrix data from pre-computed metrics."""
    metrics_file = "classifier_metrics.json"
    if not os.path.exists(metrics_file):
        return {"labels": [], "matrix": [], "accuracy": 0, "total_samples": 0}
    
    with open(metrics_file, "r") as f:
        metrics = json.load(f)
    
    labels = metrics.get("labels", [])
    report = metrics.get("classification_report", {})
    
    matrix = []
    for true_label in labels:
        row = []
        for pred_label in labels:
            if true_label == pred_label:
                support = int(report.get(true_label, {}).get("support", 1))
                recall = report.get(true_label, {}).get("recall", 0)
                row.append(int(support * recall))
            else:
                row.append(0)
        matrix.append(row)
    
    return {
        "labels": labels,
        "matrix": matrix,
        "accuracy": metrics.get("accuracy", 0),
        "total_samples": sum(sum(row) for row in matrix),
    }

def get_confusion_matrix():
    """Return confusion matrix data for the classifier."""
    metrics = classifier.get_metrics()
    
    # Build confusion matrix from classification report
    labels = metrics.get("labels", [])
    report = metrics.get("classification_report", {})
    
    matrix = []
    for true_label in labels:
        row = []
        for pred_label in labels:
            # Simplified: use precision * recall as proxy for correct/incorrect
            if true_label == pred_label:
                support = report.get(true_label, {}).get("support", 1)
                recall = report.get(true_label, {}).get("recall", 0)
                row.append(int(support * recall))
            else:
                row.append(0)
        matrix.append(row)
    
    return {
        "labels": labels,
        "matrix": matrix,
        "accuracy": metrics.get("accuracy", 0),
        "total_samples": sum(sum(row) for row in matrix),
    }

# --- Chat endpoint ---
from pydantic import BaseModel as PydanticBaseModel

class ChatRequest(PydanticBaseModel):
    question: str

@app.post("/api/chat")
def chat_with_system(request: ChatRequest):
    """Conversational chatbot that answers questions about safety data."""
    question = request.question.strip().lower()
    
    if not all_processed_reports:
        return {"answer": "No reports have been loaded yet. Please load some reports first by clicking 'Load next report'."}
    
    # Build context from all loaded reports
    total = len(all_processed_reports)
    precursors = [r for r in all_processed_reports if r.get("is_precursor")]
    non_precursors = [r for r in all_processed_reports if not r.get("is_precursor")]
    
    # Location stats
    loc_counts = {}
    loc_risk = {}
    for r in all_processed_reports:
        for loc in r["extracted_entities"].get("locations", []):
            loc_counts[loc] = loc_counts.get(loc, 0) + 1
            if loc not in loc_risk:
                loc_risk[loc] = []
            loc_risk[loc].append(r["risk_data"]["score"])
    
    # Equipment stats
    equip_counts = {}
    equip_risk = {}
    for r in all_processed_reports:
        for eq in r["extracted_entities"].get("equipment", []):
            equip_counts[eq] = equip_counts.get(eq, 0) + 1
            if eq not in equip_risk:
                equip_risk[eq] = []
            equip_risk[eq].append(r["risk_data"]["score"])
    
    # Hazard stats
    hazard_counts = {}
    for r in all_processed_reports:
        for haz in r["extracted_entities"].get("hazards", []):
            hazard_counts[haz] = hazard_counts.get(haz, 0) + 1
    
    # Build structured context
    context = {
        "total_reports": total,
        "precursors_detected": len(precursors),
        "locations": {k: {"count": v, "avg_risk": round(sum(loc_risk.get(k, [0])) / max(len(loc_risk.get(k, [1])), 1))} for k, v in sorted(loc_counts.items(), key=lambda x: x[1], reverse=True)},
        "equipment": {k: {"count": v, "avg_risk": round(sum(equip_risk.get(k, [0])) / max(len(equip_risk.get(k, [1])), 1))} for k, v in sorted(equip_counts.items(), key=lambda x: x[1], reverse=True)},
        "hazards": dict(sorted(hazard_counts.items(), key=lambda x: x[1], reverse=True)),
        "precursor_reports": [{"id": r["report"]["id"], "text": r["report"]["text"], "score": r["risk_data"]["score"], "trajectory": r["risk_data"]["trajectory"], "equipment": r["extracted_entities"].get("equipment", []), "locations": r["extracted_entities"].get("locations", [])} for r in precursors],
        "latest_report": all_processed_reports[-1]["report"] if all_processed_reports else None,
        "latest_risk": all_processed_reports[-1]["risk_data"]["score"] if all_processed_reports else 0,
    }
    
    # Try Gemini API
    api_key = os.environ.get("GEMINI_API_KEY", "")
    try:
        answer = _chat_with_gemini(question, context)
        return {"answer": answer}
    except Exception as e:
        print(f"[Chat] Gemini failed: {e}, using rule-based fallback")
    
    # Rule-based fallback
    answer = _chat_rule_based(question, context)
    return {"answer": answer}


def _chat_with_gemini(question: str, context: dict) -> str:
    api_key = os.environ.get("GEMINI_API_KEY")
    prompt = f"""You are SAFEGUARD AI, an industrial safety chatbot for Oil India Limited's Duliajan facility in Assam.

You have access to the following safety data from the facility:

{json.dumps(context, indent=2, default=str)}

A safety officer or plant manager is asking you a question. Answer it using the data above.

Rules:
- Be specific — use actual equipment names (PUMP_104, FORKLIFT, etc.), location names, and risk scores from the data
- If asked about danger zones, list the locations with highest incident counts and risk scores
- If asked about what to do next, give 2-3 actionable recommendations based on the precursor data
- If asked about most hazardous equipment or location, rank them by incident count and average risk score
- Keep answers under 100 words
- Use plain language, not jargon
- Be direct and action-oriented
- Do NOT use markdown formatting
"""
    try:
        client = google_genai.Client(api_key=api_key)
        response = client.models.generate_content(model="gemini-3.6-flash", contents=prompt)
        return response.text.strip()
    except NameError:
        import google.generativeai as legacy_genai
        legacy_genai.configure(api_key=api_key)
        model = legacy_genai.GenerativeModel("gemini-3.6-flash")
        response = model.generate_content(prompt)
        return response.text.strip()


def _chat_rule_based(question: str, context: dict) -> str:
    """Rule-based chatbot fallback when Gemini is unavailable."""
    q = question.lower()
    
    # Most hazardous location
    if any(w in q for w in ["hazardous", "danger", "dangerous", "risk", "worst", "unsafe"]):
        if "location" in q or "zone" in q or "area" in q or "place" in q:
            locs = context.get("locations", {})
            if not locs:
                return "No location data available yet."
            lines = []
            for loc, data in list(locs.items())[:5]:
                risk_level = "HIGH" if data["avg_risk"] >= 70 else "MEDIUM" if data["avg_risk"] >= 40 else "LOW"
                lines.append(f"- {loc}: {data['count']} incidents, avg risk {data['avg_risk']}/100 ({risk_level})")
            return f"The most dangerous locations at Duliajan facility are:\n" + "\n".join(lines) + "\n\nRecommendation: Increase safety inspections and deploy additional PPE at HIGH risk locations."
        
        if "equipment" in q or "machine" in q or "pump" in q:
            equips = context.get("equipment", {})
            if not equips:
                return "No equipment data available yet."
            lines = []
            for eq, data in list(equips.items())[:5]:
                risk_level = "HIGH" if data["avg_risk"] >= 70 else "MEDIUM" if data["avg_risk"] >= 40 else "LOW"
                lines.append(f"- {eq}: {data['count']} incidents, avg risk {data['avg_risk']}/100 ({risk_level})")
            return f"The most hazardous equipment identified:\n" + "\n".join(lines) + "\n\nRecommendation: Schedule immediate maintenance for HIGH risk equipment."
        
        # General most hazardous
        precursors = context.get("precursor_reports", [])
        if precursors:
            lines = []
            for p in precursors[:3]:
                lines.append(f"- {p['id']}: Risk {p['score']}/100 ({p['trajectory']}) — {p['text'][:80]}")
            return f"The most hazardous incidents detected (precursors):\n" + "\n".join(lines) + f"\n\n{context['precursors_detected']} precursor(s) detected out of {context['total_reports']} reports. Immediate intervention required."
        return f"No precursors detected yet from {context['total_reports']} reports."
    
    # What to do next
    if any(w in q for w in ["what to do", "next", "action", "recommend", "should", "prevent"]):
        precursors = context.get("precursor_reports", [])
        actions = []
        
        # Check for high-risk equipment
        equips = context.get("equipment", {})
        for eq, data in equips.items():
            if data["avg_risk"] >= 70:
                actions.append(f"URGENT: Shut down and inspect {eq} immediately (avg risk {data['avg_risk']}/100)")
            elif data["avg_risk"] >= 40:
                actions.append(f"Schedule maintenance for {eq} within 48 hours (avg risk {data['avg_risk']}/100)")
        
        # Check for high-risk locations
        locs = context.get("locations", {})
        for loc, data in locs.items():
            if data["avg_risk"] >= 70:
                actions.append(f"Restrict access to {loc} — {data['count']} incidents with high risk")
            elif data["count"] >= 3:
                actions.append(f"Increase safety patrols at {loc} ({data['count']} incidents)")
        
        # Check for unresolved precursors
        if precursors:
            actions.append(f"Investigate {len(precursors)} active precursor alert(s)")
        
        if not actions:
            actions.append("Continue routine monitoring — no immediate actions required")
            actions.append("Review closed corrective actions for effectiveness")
            actions.append("Update safety training based on recent incident patterns")
        
        return "Recommended actions based on current data:\n" + "\n".join(f"{i+1}. {a}" for i, a in enumerate(actions[:6]))
    
    # Danger zones
    if any(w in q for w in ["danger", "zone", "where", "hotspot", "cluster"]):
        locs = context.get("locations", {})
        if not locs:
            return "No location data available yet."
        
        danger = []
        watch = []
        safe = []
        for loc, data in locs.items():
            if data["avg_risk"] >= 70 or data["count"] >= 4:
                danger.append(f"RED ZONE: {loc} — {data['count']} incidents, avg risk {data['avg_risk']}/100")
            elif data["avg_risk"] >= 40 or data["count"] >= 2:
                watch.append(f"YELLOW ZONE: {loc} — {data['count']} incidents, avg risk {data['avg_risk']}/100")
            else:
                safe.append(f"{loc} — {data['count']} incident(s), low risk")
        
        result = "Facility Danger Zone Assessment:\n\n"
        if danger:
            result += "CRITICAL ZONES (immediate action):\n" + "\n".join(f"  {d}" for d in danger) + "\n\n"
        if watch:
            result += "WATCH ZONES (increased monitoring):\n" + "\n".join(f"  {w}" for w in watch) + "\n\n"
        if safe:
            result += "LOW RISK ZONES:\n" + "\n".join(f"  {s}" for s in safe[:3])
        return result
    
    # Stats / summary
    if any(w in q for w in ["summary", "stats", "overview", "status", "how many", "total"]):
        return f"Facility Safety Summary:\n- Total reports analyzed: {context['total_reports']}\n- Precursors detected: {context['precursors_detected']}\n- Unique locations: {len(context.get('locations', {}))}\n- Equipment types tracked: {len(context.get('equipment', {}))}\n- Hazard types found: {len(context.get('hazards', {}))}\n- Latest risk score: {context.get('latest_risk', 0)}/100"
    
    # Default
    return f"I can help you with safety data questions. Try asking:\n- Which location is most hazardous?\n- What are the danger zones?\n- What should I do next?\n- Give me a summary of the current status\n- Which equipment is most dangerous?\n\nCurrently analyzing {context['total_reports']} reports with {context['precursors_detected']} precursor alert(s)."



DIST_DIR = Path("frontend/dist")

STATIC_DIR = Path("static")



if DIST_DIR.exists():

    app.mount("/assets", StaticFiles(directory=DIST_DIR / "assets"), name="assets")



    @app.get("/{full_path:path}")

    async def serve_react(full_path: str):

        # Serve index.html for all non-API, non-asset routes (SPA routing)

        file_path = DIST_DIR / full_path

        if file_path.exists() and file_path.is_file():

            return FileResponse(file_path)

        return FileResponse(DIST_DIR / "index.html")

elif STATIC_DIR.exists():

    @app.get("/{full_path:path}")

    async def serve_static(full_path: str):

        file_path = STATIC_DIR / full_path

        if file_path.exists() and file_path.is_file():

            return FileResponse(file_path)

        return FileResponse(STATIC_DIR / "index.html")

else:

    @app.get("/")

    def read_root():

        return {"message": "API is running. Frontend not built yet. Run 'cd frontend && npm run build'"}





if __name__ == "__main__":

    import uvicorn

    uvicorn.run("main:app", host="127.0.0.1", port=8000)

